import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

def generate_report(df, date_col, sales_col, product_col, city_col):

    if not os.path.exists("output"):
        os.makedirs("output")

    report_path = "output/report.xlsx"

    df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
    df = df.dropna()

    total_sales = df[sales_col].sum()
    avg_sales = df[sales_col].mean()

    top_product = (
        df.groupby(product_col)[sales_col]
        .sum()
        .sort_values(ascending=False)
        .head(1)
    )

    top_product_name = top_product.index[0]
    top_product_value = top_product.iloc[0]

    sales_by_city = df.groupby(city_col)[sales_col].sum()
    monthly_sales = df.groupby(df[date_col].dt.to_period("M"))[sales_col].sum()

    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:

        summary_df = pd.DataFrame({
            "Metric": ["Total Sales", "Top Product", "Average Sale"],
            "Value": [
                f"{total_sales:,}",
                f"{top_product_name} (₹{top_product_value:,})",
                f"₹{avg_sales:.2f}"
            ]
        })

        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        sales_by_city.reset_index().to_excel(writer, sheet_name="City Sales", index=False)
        monthly_sales.reset_index().to_excel(writer, sheet_name="Monthly Sales", index=False)

    # Load workbook
    wb = load_workbook(report_path)

    # --- Monthly Chart ---
    ws = wb["Monthly Sales"]

    chart = BarChart()
    chart.title = "Monthly Sales"
    chart.style = 10
    chart.y_axis.title = "Sales"
    chart.x_axis.title = "Month"

    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    chart.dLbls = DataLabelList()
    chart.dLbls.showVal = True

    ws.add_chart(chart, "E2")

    # --- City Chart ---
    ws2 = wb["City Sales"]

    chart2 = BarChart()
    chart2.title = "Sales by City"
    chart2.style = 12
    chart2.y_axis.title = "Sales"
    chart2.x_axis.title = "City"

    data2 = Reference(ws2, min_col=2, min_row=1, max_row=ws2.max_row)
    cats2 = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)

    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)

    chart2.dLbls = DataLabelList()
    chart2.dLbls.showVal = True

    ws2.add_chart(chart2, "E2")

    wb.save(report_path)

    return report_path, total_sales, avg_sales, top_product_name, top_product_value, monthly_sales, sales_by_city